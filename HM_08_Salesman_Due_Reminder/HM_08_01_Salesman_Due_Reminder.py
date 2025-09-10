"""
🚀 HM_08_01_Salesman_Due_Reminder.py – Daily Due Reminder for Salesmen

📌 PURPOSE:
    - Find customers with DOs older than 3 days and balance > 300 BDT
    - Send personalized Excel + HTML email to each salesman
    - Send summary to management
    - Color-code: Red (>15d), Yellow (10–15d), Green (<10d)
    - Auto-delete temp files
"""

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import create_engine
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === 1. Load Environment Variables from .env ===
load_dotenv()

# Read ZID and project
try:
    ZID = int(os.environ["ZID_GULSHAN_TRADING"])
except KeyError:
    raise RuntimeError("❌ ZID_GULSHAN_TRADING not found in .env")
PROJECT_NAME = "GULSHAN TRADING"

print(f"📌 Processing for: {PROJECT_NAME} (ZID={ZID})")


# === 2. Add root (E:\) to Python path ===
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)


# === 3. Import shared modules ===
from mail import send_mail, get_email_recipients


# === 4. Create engine using DATABASE_URL ===
from project_config import DATABASE_URL, holiday
engine = create_engine(DATABASE_URL)


# === 5. Suppress warnings ===
import warnings
warnings.filterwarnings('ignore', category=pd.errors.DtypeWarning)
pd.options.mode.chained_assignment = None


# === 6. Holiday & Friday Check ===
def is_today_holiday(holiday_list):
    today = datetime.now().date().strftime("%Y-%m-%d")
    return today in holiday_list

def is_today_friday():
    return datetime.now().weekday() == 4

if is_today_holiday(holiday()):
    print("📅 Today is a holiday. Exiting.")
    sys.exit(0)

if is_today_friday():
    print("📅 Today is Friday. Exiting.")
    sys.exit(0)


# === 7. Date Setup ===
TODAY = datetime.now().strftime("%Y-%m-%d")
TWENTY_DAYS_AGO = (datetime.now() - timedelta(days=20)).strftime("%Y-%m-%d")


# === 8. Salesman Mapping (Email + Name) ===
salesman_info_dict = {
    'SA--000068': ('mssalim5904@gmail.com', 'Md. Salimullah (Salim)'),
    'SA--000224': ('aponmdarifulislam603@gmail.com', 'Md. Ariful Islam-2'),
    'SA--000038': ('hbelal706@gmail.com', 'Md. Belal Hossain'),
    'SA--000144': ('hossainmobarok45@gmail.com', 'Syed Mobarok Hossain'),
    'SA--000021': ('mdm10413@gmail.com', 'Md. Maruful Islam'),
    'SA--000193': ('limonmridha30@gmail.com', 'Limon Mridha'),
    'SA--000114': ('plabonpavel494@gmail.com', 'Md. Pavel Mia'),
    'SA--000011': ('jamal.hmbr@gmail.com', 'Jamal Hossain Titu'),
    'SA--000192': ('sajibhossen701@gmail.com', 'Sojib Hossen'),
    'SA--000098': ('javedfeni127@gmail.com', 'Md. Belayet Hossen'),
    'SA--000227': ('mdsumonhossionmithu00@gmail.com', 'Md. Sumon Hossain Mithu'),
    'SA--000242': ('iforhadul08@gmail.com', 'Md. Forhadul Islam')
}

salesman_ids = tuple(salesman_info_dict.keys())


# === 9. Fetch DOs (Deliveries) in Last 20 Days ===
query_do = f"""
    SELECT
        opdor.xdornum AS do_number,
        opdor.xcus,
        CONCAT(opdor.xcus, '--', cacus.xshort) AS customer,
        cacus.xadd1 AS address,
        cacus.xtaxnum AS mobile_number,
        cacus.xcity AS city,
        opdor.xsp,
        prmst.xname AS SName,
        opdor.xtotamt AS total_DO_amt,
        opdor.xdate AS goods_receive_date
    FROM opdor
    LEFT JOIN cacus ON opdor.xcus = cacus.xcus
    JOIN prmst ON opdor.xsp = prmst.xemp
    WHERE opdor.zid = %(zid)s
      AND cacus.zid = %(zid)s
      AND prmst.zid = %(zid)s
      AND opdor.xdate >= %(start_date)s
"""

df_do = pd.read_sql(query_do, engine, params={'zid': ZID, 'start_date': TWENTY_DAYS_AGO})
if df_do.empty:
    print("📭 No DOs found in last 20 days.")
else:
    df_do['goods_receive_date'] = pd.to_datetime(df_do['goods_receive_date']) + pd.Timedelta(days=3)
    df_do['today_date'] = TODAY
    df_do['date_diff'] = pd.to_datetime(TODAY) - pd.to_datetime(df_do['goods_receive_date'])
    df_do = df_do[df_do['date_diff'].dt.days <= 20].sort_values(['customer', 'date_diff']).drop(columns=['today_date'])


# === 10. Fetch Current Balance for These Customers ===
if not df_do.empty:
    customer_ids = tuple(df_do['xcus'].unique())
    if len(customer_ids) == 1:
        customer_ids = f"('{customer_ids[0]}')"

    query_balance = f"""
        SELECT gldetail.xsub AS xcus, SUM(gldetail.xprime) AS balance_till_today
        FROM glheader
        JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
        WHERE glheader.zid = %(zid)s
          AND gldetail.zid = %(zid)s
          AND gldetail.xproj = %(project)s
          AND gldetail.xvoucher NOT LIKE '%%OB%%'
          AND glheader.xdate <= %(today)s
          AND gldetail.xsub IN {customer_ids}
        GROUP BY gldetail.xsub
    """
    df_balance = pd.read_sql(query_balance, engine, params={'zid': ZID, 'project': PROJECT_NAME, 'today': TODAY})

    df_final = pd.merge(df_do, df_balance, on='xcus', how='left')
    df_final = df_final[df_final['balance_till_today'] > 300]
else:
    df_final = pd.DataFrame()


# === 11. Fetch Last Payment Date ===
if not df_final.empty:
    query_payment = f"""
        SELECT gldetail.xsub AS xcus, glheader.xdate AS last_payment_date, SUM(gldetail.xprime) AS last_payment
        FROM glheader
        JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
        WHERE glheader.zid = %(zid)s
          AND gldetail.zid = %(zid)s
          AND gldetail.xproj = %(project)s
          AND gldetail.xvoucher LIKE '%%RCT-%%'
          AND glheader.xdate <= %(today)s
          AND gldetail.xsub IN {customer_ids}
        GROUP BY gldetail.xsub, glheader.xdate
        ORDER BY gldetail.xsub, glheader.xdate DESC
    """
    df_payment = pd.read_sql(query_payment, engine, params={'zid': ZID, 'project': PROJECT_NAME, 'today': TODAY})
    df_payment = df_payment.drop_duplicates(subset=['xcus'], keep='first')
    df_payment['last_payment'] = df_payment['last_payment'].abs()

    df_final = pd.merge(df_final, df_payment, on='xcus', how='left')


# === 12. Final Processing ===
if not df_final.empty:
    df_final = df_final.sort_values(['xsp', 'xcus', 'date_diff'])
    df_final = df_final.iloc[:, [0, 6, 7, 2, 8, 11, 9, 10, 12, 13, 5]]  # Reorder
    df_final = df_final[df_final['date_diff'] > pd.Timedelta(days=0)]
    df_final['date_diff_days'] = df_final['date_diff'].dt.days
else:
    df_final = pd.DataFrame(columns=[
        'do_number', 'xsp', 'SName', 'customer', 'total_DO_amt',
        'goods_receive_date', 'address', 'mobile_number', 'city',
        'balance_till_today', 'last_payment_date', 'last_payment', 'date_diff_days'
    ])


# === 13. Export & Color Excel ===
OUTPUT_FILE = "HM_08_01_reminder.xlsx"
df_export = df_final.copy()
if 'date_diff_days' in df_export.columns:
    df_export = df_export.rename(columns={'date_diff_days': 'date_diff'})

df_export.to_excel(OUTPUT_FILE, index=False)

wb = load_workbook(OUTPUT_FILE)
ws = wb.active
col_idx = 9  # date_diff column (J)

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=col_idx)
    try:
        days = int(cell.value)
        if days > 15:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        elif days >= 10:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        else:
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    except:
        pass

wb.save(OUTPUT_FILE)
print(f"✅ Colored Excel saved: {OUTPUT_FILE}")


# === 14. Send Emails to Salesmen ===
for sp_id, (email, name) in salesman_info_dict.items():
    sp_df = df_final[df_final['xsp'] == sp_id]
    if not sp_df.empty:
        file_path = f"salesman_{sp_id}.xlsx"
        sp_df.to_excel(file_path, index=False)
        html_body = [(sp_df, f"Customer Due Balance >300 BDT (Last 20 Days)")]
        send_mail(
            subject=f"HM_08.1 Customer Due Reminder – {name}",
            bodyText=f"Dear {name},<br><br>Please collect due from the following customers.",
            attachment=[file_path],
            recipient=[email],
            html_body=html_body
        )
        print(f"📨 Sent to {name} ({email})")


# === 15. Send Summary to Management ===
try:
    recipients = get_email_recipients(os.path.splitext(os.path.basename(__file__))[0])
    print(f"📬 Recipients: {recipients}")
except Exception as e:
    print(f"⚠️ Failed to get recipients: {e}. Using fallback.")
    recipients = ["ithmbrbd@gmail.com"]

send_mail(
    subject="HM_08.1 Summary – Customer Due Reminder to Salesmen",
    bodyText=f"Daily due reminder sent to {len([k for k, v in salesman_info_dict.items() if not df_final[df_final['xsp']==k].empty])} salesmen. See summary.",
    attachment=[OUTPUT_FILE],
    recipient=recipients,
    html_body=[(df_final, "All Customers with Due >300 BDT")]
)


# === 16. Cleanup Temp Files ===
import glob
for file in glob.glob("salesman_*.xlsx"):
    try:
        os.remove(file)
        print(f"🗑️ Deleted: {file}")
    except Exception as e:
        print(f"❌ Failed to delete {file}: {e}")


# === 17. Cleanup ===
engine.dispose()