"""
🚀 HM_01_Acct_Rec_Pay.py – Accounts Receivable & Payable Report

📌 PURPOSE:
    Generates AR and AP reports for multiple HMBR companies.
    - accountsReceivable.xlsx: AR per company (with 2-month comparison)
    - accountsPayable.xlsx: AP per company
    - Sends HTML email summaries with Excel attachments

🔧 DATA SOURCES:
    - GL: glheader, gldetail
    - Masters: cacus (customer), casup (supplier)
    - Database: PostgreSQL via DATABASE_URL in project_config.py

🏢 COMPANIES:
    Dynamically loaded from .env: PROJECT_100000=Karigor Ltd., etc.

📧 EMAIL:
    - Receivables: get_email_recipients("HM_01_Acct_Rec")
    - Payables: get_email_recipients("HM_01_Acct_Pay")
    - Fallback: fallback_email@example.com

💡 NOTE:
    - Uses parameterized queries to prevent SQL injection.
    - Project names are dynamic from .env → rename safely.
    - Maintains original logic: 24-month AR, 1-month AP, Excel sum row.
"""

import sys
import os
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from sqlalchemy import create_engine
from dotenv import load_dotenv
import openpyxl


# === 1. Load Environment Variables from .env ===
load_dotenv()

# Build project_zid from .env: PROJECT_100000=Karigor Ltd.
project_zid = {}
for key, value in os.environ.items():
    if key.startswith("PROJECT_"):
        try:
            zid = int(key.replace("PROJECT_", ""))
            project_zid[value.strip()] = zid
        except ValueError:
            print(f"⚠️ Invalid ZID in .env: {key}={value}")

if not project_zid:
    raise RuntimeError("❌ No valid PROJECT_* entries found in .env file. Please check.")

# Reverse mapping: zid → project name
zid_to_name = {v: k for k, v in project_zid.items()}


# === 2. Add root (E:\) to Python path ===
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

# === 3. Import shared modules ===
from mail import send_mail, get_email_recipients
from project_config import DATABASE_URL

# === 4. Create engine using shared DATABASE_URL ===
engine = create_engine(DATABASE_URL)

# === 5. Suppress warnings ===
import warnings
warnings.filterwarnings('ignore', category=pd.errors.DtypeWarning)


# === 6. Date Setup ===
this_datetime = datetime.now()
end_year = this_datetime.year
end_month = this_datetime.month
last_year = (this_datetime - relativedelta(months=1)).year
last_month = (this_datetime - relativedelta(months=1)).month

# Month labels for columns
month_list_24 = [(this_datetime - relativedelta(months=i)).strftime('%Y/%m') for i in range(24)]


# === 7. SQL Query Functions (Parameterized) ===
def get_acc_receivable(zid, project, year, month):
    """Fetch customer-wise AR up to given year/month."""
    year_month = f"{year}{month:02d}"
    query = """
        SELECT 
            gldetail.xsub,
            cacus.xorg,
            cacus.xadd2,
            cacus.xcity,
            cacus.xstate,
            SUM(gldetail.xprime) AS ar
        FROM glheader
        JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
        JOIN cacus ON gldetail.xsub = cacus.xcus
        WHERE glheader.zid = %(zid)s
          AND gldetail.zid = %(zid)s
          AND cacus.zid = %(zid)s
          AND gldetail.xproj = %(project)s
          AND gldetail.xvoucher NOT LIKE 'OB-%%'
          AND CONCAT(glheader.xyear, LPAD(glheader.xper::text, 2, '0')) <= %(year_month)s
        GROUP BY gldetail.xsub, cacus.xorg, cacus.xadd2, cacus.xcity, cacus.xstate
    """
    return pd.read_sql(query, engine, params={
        'zid': zid,
        'project': project,
        'year_month': year_month
    })


def get_acc_payable(zid, project, year, month):
    """Fetch supplier-wise AP up to given year/month."""
    year_month = f"{year}{month:02d}"
    query = """
        SELECT 
            gldetail.xsub,
            casup.xorg,
            SUM(gldetail.xprime) AS ap
        FROM glheader
        JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
        JOIN casup ON gldetail.xsub = casup.xsup
        WHERE glheader.zid = %(zid)s
          AND gldetail.zid = %(zid)s
          AND casup.zid = %(zid)s
          AND gldetail.xproj = %(project)s
          AND gldetail.xvoucher NOT LIKE 'OB-%%'
          AND CONCAT(glheader.xyear, LPAD(glheader.xper::text, 2, '0')) <= %(year_month)s
        GROUP BY gldetail.xsub, casup.xorg
    """
    return pd.read_sql(query, engine, params={
        'zid': zid,
        'project': project,
        'year_month': year_month
    })


# === 8. Accounts Receivable Processing ===
ar_dataframes = {}
ar_summaries = {}

for project, zid in project_zid.items():
    print(f"📊 Processing AR: {project} (zid={zid})")

    # Fetch current and previous month data
    df_curr = get_acc_receivable(zid, project, end_year, end_month)
    df_prev = get_acc_receivable(zid, project, last_year, last_month)

    if df_curr.empty:
        print(f"  ⚠️ No AR data for {project}")
        continue

    # Rename columns BEFORE selecting
    df_curr = df_curr.rename(columns={'xsub': 'Code', 'ar': month_list_24[0]})
    df_prev = df_prev.rename(columns={'xsub': 'Code', 'ar': 'ar_prev'})

    # Handle empty df_prev safely
    prev_subset = df_prev[['Code', 'ar_prev']] if not df_prev.empty else pd.DataFrame(columns=['Code', 'ar_prev'])

    # Merge current with previous
    df_final = df_curr.merge(prev_subset, on='Code', how='left')

    # Fill NaN and finalize column names
    df_final[month_list_24[0]] = df_final[month_list_24[0]].fillna(0)
    df_final['ar_prev'] = df_final['ar_prev'].fillna(0)
    df_final = df_final.rename(columns={
        'xorg': 'Name',
        'xadd2': 'Address',
        'xcity': 'City',
        'xstate': 'Market',
        'ar_prev': month_list_24[1]
    })

    ar_dataframes[project] = df_final

    # Summary by City
    summary = df_final.groupby('City')[[month_list_24[0], month_list_24[1]]].sum().round(2).reset_index()
    ar_summaries[project] = summary


# === 9. Export Receivables to Excel ===
ar_excel_file = 'accountsReceivable.xlsx'
with pd.ExcelWriter(ar_excel_file, engine='openpyxl') as writer:
    for project, df in ar_dataframes.items():
        safe_name = "".join(c for c in project if c.isalnum() or c in " _-")[:31]
        df.to_excel(writer, sheet_name=safe_name, index=False)

# Add sum row in first sheet using openpyxl
wb = openpyxl.load_workbook(ar_excel_file)
ws = wb.active  # First sheet

# Insert row at top
ws.insert_rows(0)

# Sum column G (current month AR) from row 3 onward
col_index = 7  # G = 7
total = sum(ws.cell(row=r, column=col_index).value or 0 for r in range(3, ws.max_row + 1))
ws.cell(row=1, column=col_index, value=total)

wb.save(ar_excel_file)


# === 10. Send Receivable Email ===

try:
    # Extract report name from filename
    report_name = os.path.splitext(os.path.basename(__file__))[0]
    recipients = get_email_recipients('HM_01_Acct_Rec')
    print(f"📬 Recipients: {recipients}")
except Exception as e:
    print(f"⚠️ Failed to fetch recipients: {e}")
    recipients = ["ithmbrbd@gmail.com"]  # Fallback


# Prepare HTML tables
html_df_list_rec = [
    (summary_df, f"{project} Summary")
    for project, summary_df in ar_summaries.items()
]

send_mail(
    subject="HM_01 Accounts Receivable Details",
    bodyText="Please find the receivable summary below.",
    attachment=[ar_excel_file],
    recipient=recipients,
    html_body=html_df_list_rec
)


# === 11. Accounts Payable Processing ===
ap_dataframes = {}

for project, zid in project_zid.items():
    print(f"💼 Processing AP: {project} (zid={zid})")
    df = get_acc_payable(zid, project, end_year, end_month)
    if not df.empty:
        df = df.rename(columns={'xsub': 'Code', 'xorg': 'Name', 'ap': 'AP'})
    else:
        df = pd.DataFrame(columns=['Code', 'Name', 'AP'])
    ap_dataframes[project] = df


# === 12. Export Payables to Excel ===
ap_excel_file = 'accountsPayable.xlsx'
with pd.ExcelWriter(ap_excel_file, engine='openpyxl') as writer:
    for project, df in ap_dataframes.items():
        safe_name = "".join(c for c in project if c.isalnum() or c in " _-")[:31]
        df.to_excel(writer, sheet_name=safe_name, index=False)


# === 13. Send Payable Email ===
try:
    recipients = get_email_recipients("HM_01_Acct_Pay")
    print(f"📬 Payable recipients: {recipients}")
except Exception as e:
    print(f"⚠️ Fallback: {e}")
    recipients = ["ithmbrbd@gmail.com"]

html_df_list_pay = [
    (df, f"{project} Payable")
    for project, df in ap_dataframes.items()
    if len(df) > 0
]

send_mail(
    subject="HM_01 Accounts Payable Details",
    bodyText="Please find the payable summary below.",
    attachment=[ap_excel_file],
    recipient=recipients,
    html_body=html_df_list_pay
)


# === 14. Cleanup ===
engine.dispose()